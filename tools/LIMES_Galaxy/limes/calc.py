
from __future__ import annotations

"""
Définit les classes suivantes :

    Reader_csv()
    Reader_excel()

Ces deux classes héritent de Source ; leur méthode .load() charge le contenu
d'un fichier CSV ou Excel
"""

from itertools import count
from .core import (Echantillon,Methode,Source,get_text,RedundantNameError,
                   CodeEspeceType,Espace)
from contextlib import closing,nullcontext

from typing import (Set,Optional,Tuple,List,Union,TextIO,cast,
                    Iterator,TYPE_CHECKING,Sequence)

# -- Lecture des fichiers sources ------------

LigneType=Sequence
# Nécessiterait la définition d'un Generic pour assurer la cohérence entre
# toutes les fonctions échangeant des LigneType.

"""
_calc_Reader(redond)

Reader générique pour les fichiers de type Excel. Les classes dérivées sont
spécialisées dans les fichiers CSV ou XLS/XLSX.

Si 'redond vaut False, _run() génère une exception RedundantNameError si deux
méthodes ou deux échantillons portent le même nom. Si 'redond vaut True, les
noms redondants sont augmentés d'un suffixe numérique pour les rendre uniques.

La classe dérivée doit définir les méthodes suivantes :

    load()  C'est la méthode lancée par l'utilisateur. Lit l'ensemble du
            fichier, renseigne l'instance et rend la liste des Methode. Cette
            fonction doit ouvrir le contexte et appeler ._run() pour effectuer
            le traitement.

    _getlg() Rend un itérateur retournant une ligne à chaque appel. La nature
            des lignes retournées est quelconque, mais ce seront ces lignes qui
            seront passées en argument aux autres fonctions spécifiques.

    _cells(lg)
            Rend un itérateur retournant la valeur de chacune des cellules de
            la ligne 'lg. Les valeurs rendues doivent être de type str.

    _cell(lg,ind,fmt)
            Rend la valeur de la cellule à l'indice 'ind de la ligne 'lg. Si
            'fmt vaut True, la valeur rendue doit être de type str (tête de
            ligne). Si 'fmt vaut False, la valeur peut être str ou int.

    _nonvide(lg)
            Rend False si le ligne 'lg est vide, True si elle comprend au moins
            une cellule non vide.

    _rewind()
            La lecture du fichier s'effectue en deux passes. Cette méthode est
            appelée avant la 2ième passe pour réinitialiser le reader si
            nécessaire.

Le schéma d'utilisation est le suivant. Pour un classe dérivée XX :
    - L'utilisateur crée l'instance de la classe XX en fonction du type de
        fichier, avec les arguments spécifiques.
    - Il appelle instance.load().
    - La méthode .load() initialise le contexte (ouvre la source) et appelle
        self._run().
    - Celle-ci est le moteur de lecture implémenté dans _calc_Reader. Il
        appelle les diverses fonctions de bases ._getlg(), ._cells(), ._cell(),
        ._nonvide(), ._rewind(), développées spécifiquement par XX, et affecte
        les attributs .methodes et .echantillons.
    - .load() affecte les différents attributs de l'instance, spécifique au
        type, rend la liste des Methode.

La seule méthode publique des classes dérivées de _calc_Reader est donc
.load(), qui rend une liste : chaque élément est une instance de Methode,
correspondant à une colonne du tableau, ordonnées de gauche à droite.
"""
class _calc_Reader(Source):

    marqueur="LIMES"
##    echantillons: List[Echantillon]

    def __init__(self,redond: bool):
        self.__redond=redond

    """
    'vus est l'ensemble des noms déjà utilisés. Si 'nom existe dans 'vus,
    génère un nouveau nom en lui accolant un suffixe numérique. Dans tous les
    cas, rend le nom utile ('nom ou 'nom_N) après l'avoir ajouté dans 'vus.
    """
    def __noredond(self,nom: str,vus: Set[str]) -> str:
        if nom in vus:
            if not self.__redond:
                raise RedundantNameError(
                    get_text("Nom de méthode ou d'échantillon <%s> redondant",
                             "Redundant name of method or sample <%s>")%
                    nom)
            for i in count(start=2):
                tt="%s_%d"%(nom,i)
                if tt not in vus:
                    nom=tt
                    break
        vus.add(nom)
        return nom

    """
    Si le marqueur existe dans la ligne 'lg, rend son indice. Sinon, rend
    None.
    """
    def __in_marqueur(self,lg: LigneType) -> Optional[int]:
        for i,v in enumerate(self._cells(lg)):
            if v==self.marqueur:
                return i
        return None

    """
    'lg est la ligne d'en-tête. Extrait les titres de colonnes et leur
    indice (il peut y avoir des colonnes vides). Les titres de colonnes sont
    extraits à partir de l'indice 'first. Rend un couple (a,b), où 'a est la
    liste des titres et 'b la liste, de même longueur, de l'indice correspon-
    dant.
    """
    def __mkcols(self,lg: LigneType,first: int) -> Tuple[List[str],List[int]]:
        tt=[]
        ind=[]
        for i,t in enumerate(self._cells(lg)):
            if i>=first and t:
                tt.append(t)
                ind.append(i)
        return (tt,ind)

    """
    'lg est la première ligne de données. Rend l'indice du mot le plus à droite
    avant l'indice 'end. Rend None si la ligne est vide avant 'end.
    """
    def __mkcol0(self,lg: LigneType,end: int) -> Optional[int]:
        ind=None
        for i,t in enumerate(self._cells(lg)):
            if i==end:
                break
            if t:
                ind=i
        return ind

    """
    Lit l'ensemble du fichier et affecte .methodes et .echantillons. Génère une
    exception si problème.
    """
    def _run(self) -> None:
        ett=lg1=None
        idx_marqueur=None
        for numlg,lg in enumerate(self._getlg()):
            idx_marqueur=self.__in_marqueur(lg)
            if idx_marqueur is not None:
                ett=lg1=lg
                numett=numlg
                break
            else:
                if self._nonvide(lg):
                    if ett is None:
                        ett=lg
                        numett=numlg
                    elif lg1 is None:
                        lg1=lg
        # A partir d'ici :
        # idx_marqueur  Numéro de colonne du marqueur "LIMES", ou None si pas
        #               de marqueur.
        # ett           Ligne donnant les noms des méthodes.
        # numett        Numéro de cette ligne.
        # lg1           Première ligne de données si 'idx_marqueur vaut None,
        #               égale à 'ett sinon.
        
        if ett is None or lg1 is None:
            raise Exception(get_text("Fichier vide","Empty file"))
            # Attention, le fichier peut encore être vide si le marqueur
            # a été fourni, car dans ce cas 'lg1 n'est pas significatif (on n'a
            # pas lu encore de ligne de donnée).

        if idx_marqueur is None:
            titres,cols=self.__mkcols(ett,1)
            ok=len(titres)>0
            if ok:
                col0=self.__mkcol0(lg1,cols[0])
                ok=col0 is not None
        else:
            col0=idx_marqueur
            titres,cols=self.__mkcols(ett,col0+1)
            ok=len(titres)>0
        if not ok:
            raise Exception(get_text("Fichier mal formé","Badly formated file"))
            
        # A partir d'ici :
        # col0      Numéro de la colonne donnant les noms des échantillons.
        # titres    Liste des noms de méthodes.
        # cols      Liste des numéros des colonnes de méthodes.
        assert col0 is not None

        codesespeces: List[List[Optional[CodeEspeceType]]] =[[] for _ in titres]
        # Attention, ne pas remplacer par [[]]*len(titres) !
        self.echantillons=[]
        self._rewind()
        s:Set[str] =set()
        vu=False
        for numlg,lg in enumerate(self._getlg()):
            if numlg>numett:
                ech=self._cell(lg,col0,True)
                assert isinstance(ech,str)
                if ech:
                    self.echantillons.append(Echantillon(self.__noredond(ech,s)))
                    for e,esp in zip(cols,codesespeces):
                        v: Union[int,str,None] =self._cell(lg,e,False)
                        if not v:
                            raise Exception(get_text("Cellule vide (%dx%d)",
                                                     "Empty cell (%dx%d)")%
                                            (numlg+1,e+1))
                        if v in ("-","?"): v=None
                        esp.append(v)
                    vu=True
        if not vu:
            raise Exception(get_text("Fichier vide","Empty file"))
            # Ceci garantit que le nombre d'échantillons des méthodes produites
            # est >=1.
        s=set()
        self.methodes=[]
        for titre,esp in zip(titres,codesespeces):
            ll=[(ech,cd) for ech,cd in zip(self.echantillons,esp)
                if cd is not None]
            lech,lcd=zip(*ll)
            self.methodes.append(Methode(self.__noredond(titre,s),lech,lcd,
                                         self))

    if TYPE_CHECKING:
        def _getlg(self) -> Iterator[LigneType]:
            yield range(0) # un Iterable quelconque.

        def _cells(self,lg: LigneType) -> Iterator[str]:
            yield ""

        def _cell(self,lg: LigneType,ind: int,fmt: bool) -> Union[str,int]:
            return 0

        def _nonvide(self,lg: LigneType) -> bool:
            return False

    def _rewind(self) -> None:
        pass

"""
Reader_csv(fich,separ=',',redond=False)

Reader pour le fichier 'fich de format CSV. 'separ est le caractère séparateur.

La lecture s'effectue par .load(). Chaque colonne correspond à une méthode.

Deux formats sont acceptés :
1. La case supérieure gauche du tableau doit contenir le mot "DATA". La ligne
correspondante donne les noms de méthodes, la colonne les noms d'échantillons.
Les lignes au-dessus de la ligne DATA, et les colonnes à gauche, sont ignorées.
2. Pas de marqueur "DATA". Dans ce cas, la première ligne non vide représente
la ligne de titres. La colonne non vide la plus à droite (avant la première
colonne de titre) de la première ligne de données est identifiée comme la
colonne donnant les noms des échantillons. Les colonnes à gauche de celle-ci
sont ignorées.

Il peut y avoir des colonnes vides entre les colonnes de méthodes, ou des
lignes vides entre les lignes d'échantillons. Les lignes pour lesquelles la
colonne du nom de l'échantillon est vide sont également ignorées.

Si 'redond vaut False, load() génère une exception RedundantNameError si deux
méthodes ou deux échantillons portent le même nom. Si 'redond vaut True, les
noms redondants sont augmentés d'un suffixe numérique ("nom_N") pour les rendre
uniques. (note : il pourrait être utile de retourner en résultat cette
information).

Les codes-espèces sont de type str ou int. Les échantillons non pris en compte
par une méthode doivent être marqués '-' ou '?' : ainsi, toutes les méthodes
n'ont pas forcément le même nombre d'échantillons.

L'instance dispose des attributs et méthodes suivants, en plus de ceux hérités
de Source :
    .type       "csv".
    .separ      Le caractère séparateur.
"""
class Reader_csv(_calc_Reader):
    type="csv"

    def __init__(self,fich: str,separ: str=',',redond: bool=False):
        _calc_Reader.__init__(self,redond)
        self.fich=fich
        self.separ=separ
        import csv
        self.__csv=csv

    """
    Procède à la lecture du fichier, et rend une liste d'instances de Methode
    extraites du fichier CSV, dans l'ordre du fichier source, de gauche à
    droite.
    Pour toutes les méthodes, les échantillons sont ordonnés dans l'ordre du
    fichier source, de haut en bas. De plus, un échantillon est représenté par
    la même instance de Echantillon pour toutes les Methode.
    Génère une exception si le fichier n'existe pas, ne contient aucune colonne
    de méthodes ou aucune ligne d'échantillons, ou est mal formé.
    """
    def load(self) -> List[Methode]:
        if hasattr(self,"methodes"):
            return self.methodes
        f=open(self.fich,newline='')
        try:
            with f:
                self.f=f
                self.reader=self.__csv.reader(f,delimiter=self.separ)
                self._run()
                return self.methodes
        except Exception as e:
            raise type(e)(get_text("Ne peut charger le fichier CSV\n%s",
                                   "Cannot load the CSV file %s\n")%self.fich)

    def _getlg(self) -> Iterator[LigneType]:
        yield from self.reader

    def _cells(self,lg: LigneType) -> Iterator[str]:
        yield from lg

    def _cell(self,lg: LigneType,ind: int,fmt: bool) -> str:
        return lg[ind]

    _nonvide=any # type: ignore

    def _rewind(self):
        self.f.seek(0)
        self.reader=self.__csv.reader(self.f,delimiter=self.separ)
        # Nécessaire ?

"""
Reader_excel(fich,feuille=0,redond=False)

Comme Reader_csv(), mais traite un fichier Excel de format .xls. 'feuille
identifie la feuille dans le classeur, soit par son numéro (à compter de 0),
soit par son nom ; par défaut, traite la 1ière feuille.

L'instance dispose des attributs et méthodes suivants, en plus de ceux hérités
de Source :
    .type       "excel".
    .feuille    Affecté initialement au paramètre 'feuille, puis réaffecté
                après le chargement au nom réel de la feuille.
    .get_sheets()
"""
class Reader_excel(_calc_Reader):
    type="excel"
    if TYPE_CHECKING:
        import xlrd
        module=xlrd
        # Nécessaire pour forcer le type checking !
    else:
        module=None

    @staticmethod
    def __import() -> None:
        if Reader_excel.module is None:
            import xlrd
            Reader_excel.module=xlrd

    """
    Rend la liste des feuilles du fichier 'fich. Staticmethod.
    """
    @classmethod
    def get_sheets(cls,fich: str) -> List[str]:
        cls.__import()
        with cls.module.open_workbook(fich,on_demand=True) as f:
            return f.sheet_names()

    def __init__(self,fich: str,feuille: Union[int,str] =0,redond: bool=False):
        super().__init__(redond)
        self.fich=fich
        self.feuille=feuille
        self.__import()

    """
    Procède à la lecture du fichier, et rend la liste d'instances de Methode
    extraites du fichier Excel.
    Pour toutes les méthodes, les échantillons sont ordonnés dans l'ordre du
    fichier source, de haut en bas. De plus, un échantillon est représenté par
    la même instance de Echantillon pour toutes les Methode.
    Génère une exception si le fichier n'existe pas, ne contient aucune colonne
    de méthodes ou aucune ligne d'échantillons, ou est mal formé.
    """
    def load(self) -> List[Methode]:
        if hasattr(self,"methodes"):
            return self.methodes
        try:
            with self.module.open_workbook(self.fich,on_demand=True) as f:
                self.sheet=f.sheet_by_index(self.feuille) \
                                    if isinstance(self.feuille,int) \
                                    else f.sheet_by_name(self.feuille)
                try:
                    self._run()
                    self.feuille=self.sheet.name
                    return self.methodes
                finally:
                    f.unload_sheet(self.feuille)
                    # Nécessaire ? On ne sait pas si le contexte de 'f libère toutes
                    # les ressources, y compris celles du sheet.
        except Exception as e:
            raise type(e)(get_text("Ne peut charger le fichier Excel\n%s",
                                   "Cannot load the Excel file %s\n")%
                          self.fich)

    def _getlg(self) -> Iterator[LigneType]:
        for i in range(self.sheet.nrows):
            yield self.sheet.row(i)

    def _cells(self,lg: LigneType) -> Iterator[str]:
        for c in lg:
            yield str(c.value)

    def _cell(self,lg: LigneType,ind: int,fmt: bool) -> Union[str,int]:
        c=lg[ind]
        if c.ctype==self.module.XL_CELL_DATE:
            raise Exception(get_text("Cellule de type date",
                                     "Date-type cell"))
        v=c.value
        if not isinstance(v,str):
            v=int(v)
            if fmt: v=str(v)
        return v

    def _nonvide(self,lg: LigneType) -> bool:
        for c in lg:
            if c.ctype!=self.module.XL_CELL_EMPTY:
                return True
        return False

from typing import Type

"""
Reader_excelx(fich,feuille=0,redond=False)

Comme Reader_excel(), mais traite un fichier Excel de format .xlsx.
"""
class Reader_excelx(_calc_Reader):
    type="excelx"
    if TYPE_CHECKING:
        import openpyxl
        module=openpyxl
    else:
        module=None

    @staticmethod
    def __import() -> None:
        if Reader_excelx.module is None:
            import openpyxl
            Reader_excelx.module=openpyxl

    @classmethod
    def get_sheets(cls,fich: str) -> List[str]:
        cls.__import()
        with closing(cls.module.load_workbook(fich,read_only=True)) as f:
            return f.sheetnames

    def __init__(self,fich: str,feuille: Union[int,str] =0,redond: bool=False):
        super().__init__(redond)
        self.fich=fich
        self.feuille=feuille
        self.__import()

    """
    Procède à la lecture du fichier, et rend la liste d'instances de Methode
    extraites du fichier Excel.
    Pour toutes les méthodes, les échantillons sont ordonnés dans l'ordre du
    fichier source, de haut en bas. De plus, un échantillon est représenté par
    la même instance de Echantillon pour toutes les Methode.
    Génère une exception si le fichier n'existe pas, ne contient aucune colonne
    de méthodes ou aucune ligne d'échantillons, ou est mal formé.
    """
    def load(self) -> List[Methode]:
        if hasattr(self,"methodes"):
            return self.methodes
        try:
            with closing(self.module.load_workbook(self.fich,
                                                   read_only=True)) as f:
                self.sheet=f.worksheets[self.feuille] \
                                    if isinstance(self.feuille,int) \
                                    else f[self.feuille]
                self._run()
                self.feuille=self.sheet.title
                return self.methodes
        except Exception as e:
            raise type(e)(get_text("Ne peut charger le fichier Excel\n%s",
                                   "Cannot load the Excel file %s\n")%
                          self.fich)

    def _getlg(self) -> Iterator[LigneType]:
        yield from self.sheet.iter_rows(values_only=True)

    def _cells(self,lg: LigneType) -> Iterator[str]:
        yield from lg

    def _cell(self,lg: LigneType,ind: int,fmt: bool) -> Union[str,int]:
        import datetime
        v=lg[ind]
        if v is None:
            v=""
        elif isinstance(v,datetime.datetime):
            raise Exception(get_text("Cellule de type date","Date-type cell"))
        elif not isinstance(v,str) and fmt:
            v=str(v)
        return v

    def _nonvide(self,lg: LigneType) -> bool:
        for c in lg:
            if c is not None:
                return True
        return False

"""
Sauvegarde l'Espace espace dans le fichier 'fich. 'separ est le séparateur de
champs. 'fich peut être un objet file-like déjà ouvert.
"""
def Writer_csv(fich: Union[TextIO,str],espace: Espace,separ: str =','):
    if isinstance(fich,str):
        cm=open(fich,"w")
    else:
        cm=cast(TextIO,nullcontext(fich))
    with cm as f:
        f.write(Reader_csv.marqueur)
        for m in espace:
            f.write("%s%s"%(separ,m.nom))
        f.write("\n")
        for e in espace.echantillons:
            f.write(e.nom)
            for m in espace:
                f.write("%s%s"%(separ,str(m.get(e,"-"))))
            f.write("\n")
